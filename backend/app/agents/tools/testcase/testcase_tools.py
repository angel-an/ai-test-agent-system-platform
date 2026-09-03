"""
测试用例管理工具

提供测试用例创建、更新和批量操作的 HTTP 接口调用工具。
"""
"""
andan
"""


import logging
from types import SimpleNamespace
from typing import Optional, Any

import httpx
from langchain_core.tools import tool

from app.config.settings import settings

logger = logging.getLogger(__name__)

API_BASE_URL = f"http://localhost:{settings.app_port}"
API_PREFIX = settings.api_prefix

# type: ignore  MC80OmFIVnBZMlhscm9ua3VMazZiWGREVlE9PTphNTIyZWE4MA==

def _get_api_url(path: str) -> str:
    """构建完整的 API URL"""
    return f"{API_BASE_URL}{API_PREFIX}{path}"


async def _make_http_request(
    method: str,
    url: str,
    json_data: Optional[dict] = None,
    params: Optional[dict] = None,
) -> dict[str, Any]:
    """发送 HTTP 请求的通用函数"""
    try:
        async with httpx.AsyncClient(timeout=30.0) as client:
            response = await client.request(
                method=method,
                url=url,
                json=json_data,
                params=params,
            )
            response.raise_for_status()
            return response.json()
    except httpx.HTTPStatusError as e:
        error_detail = e.response.text
        try:
            error_json = e.response.json()
            error_detail = error_json.get("detail", error_detail)
        except Exception:
            pass
        raise Exception(f"HTTP {e.response.status_code}: {error_detail}")
    except httpx.RequestError as e:
        raise Exception(f"网络请求失败: {str(e)}")
    except Exception as e:
        raise Exception(f"请求失败: {str(e)}")


async def _create_test_case_impl(
    project_identifier: str,
    folder_id: str,
    name: str,
    description: Optional[str] = None,
    preconditions: Optional[str] = None,
    priority: str = "medium",
    status: str = "new",
    case_type: str = "functional",
    owner: Optional[str] = None,
    tags: Optional[list[str]] = None,
    issues: Optional[list[str]] = None,
    automation_status: str = "not_automated",
    custom_fields: Optional[dict[str, Any]] = None,
    template: str = "test_case",
    test_case_steps: Optional[list[dict[str, str]]] = None,
    feature: Optional[str] = None,
    scenario: Optional[str] = None,
    background: Optional[str] = None,
    module: Optional[str] = None,
    keyword: Optional[str] = None,
    tp_code: Optional[str] = None,
    risk_level: Optional[str] = None,
    case_kind: Optional[str] = "sit",
) -> dict[str, Any]:
    """创建测试用例的内部实现"""
    request_data: dict[str, Any] = {
        "name": name,
        "template": template,
        "priority": priority,
        "status": status,
        "case_type": case_type,
        "automation_status": automation_status,
    }
# pragma: no cover  MS80OmFIVnBZMlhscm9ua3VMazZiWGREVlE9PTphNTIyZWE4MA==

    if description is not None:
        request_data["description"] = description
    if preconditions is not None:
        request_data["preconditions"] = preconditions
    if owner is not None:
        request_data["owner"] = owner
    if tags is not None:
        request_data["tags"] = tags
    if issues is not None:
        request_data["issues"] = issues
    if custom_fields is not None:
        request_data["custom_fields"] = custom_fields
    if module is not None:
        request_data["module"] = module
    if keyword is not None:
        request_data["keyword"] = keyword
    if tp_code is not None:
        request_data["tp_code"] = tp_code
    if risk_level is not None:
        request_data["risk_level"] = risk_level
    if case_kind is not None:
        request_data["case_kind"] = case_kind

    if template == "test_case_bdd":
        if feature is not None:
            request_data["feature"] = feature
        if scenario is not None:
            request_data["scenario"] = scenario
        if background is not None:
            request_data["background"] = background
    else:
        if test_case_steps is not None:
            request_data["test_case_steps"] = test_case_steps

    url = _get_api_url(f"/projects/{project_identifier}/folders/{folder_id}/test-cases")
    response_data = await _make_http_request(method="POST", url=url, json_data=request_data)

    if response_data.get("success"):
        test_case_data = response_data.get("data", {})
        return {
            "success": True,
            "data": test_case_data,
            "message": f"测试用例 {test_case_data.get('identifier', '')} 创建成功"
        }
    else:
        return {"success": False, "error": "API 返回失败", "message": "创建测试用例失败"}


@tool
async def create_test_case_tool(
    project_identifier: str,
    folder_id: str,
    name: str,
    description: Optional[str] = None,
    preconditions: Optional[str] = None,
    priority: str = "medium",
    status: str = "new",
    case_type: str = "functional",
    owner: Optional[str] = None,
    tags: Optional[list[str]] = None,
    issues: Optional[list[str]] = None,
    automation_status: str = "not_automated",
    custom_fields: Optional[dict[str, Any]] = None,
    template: str = "test_case",
    test_case_steps: Optional[list[dict[str, str]]] = None,
    feature: Optional[str] = None,
    scenario: Optional[str] = None,
    background: Optional[str] = None,
    module: Optional[str] = None,
    keyword: Optional[str] = None,
    tp_code: Optional[str] = None,
    risk_level: Optional[str] = None,
    case_kind: Optional[str] = "sit",
) -> dict[str, Any]:
    """
    创建测试用例工具（通过 HTTP 接口调用）。

    支持普通测试用例和 BDD 测试用例两种模板。

    Args:
        project_identifier: 项目标识符，如 'PROJ-001'
        folder_id: 文件夹 UUID
        name: 测试用例名称（必填）
        description: 测试用例描述（可选，支持 HTML）
        preconditions: 前置条件（可选，支持 HTML）
        priority: 优先级，可选值：critical, high, medium, low（默认 medium）
        status: 状态，默认 new
        case_type: 测试类型，默认 functional
        owner: 负责人邮箱（可选）
        tags: 标签列表（可选）
        issues: 关联的 Jira issues（可选）
        automation_status: 自动化状态，默认 not_automated
        custom_fields: 自定义字段（可选）
        template: 模板类型，默认 test_case
        test_case_steps: 测试步骤列表（普通测试用例使用）
        feature: BDD Feature 描述（BDD 测试用例必填）
        scenario: BDD Scenario 描述（BDD 测试用例必填）
        background: BDD Background 描述（BDD 测试用例可选）
        module: 所属模块（可选）
        keyword: 关键词，可选值：正向, 反向, 边界（可选）
        tp_code: 测试点编号，如 TP-001（可选）
        risk_level: 风险等级，可选值：high, medium, low（可选）
        case_kind: 用例种类，可选值：sit, uat, smoke, api（默认 sit）

    Returns:
        dict: 包含创建结果的字典
    """
    try:
        return await _create_test_case_impl(
            project_identifier=project_identifier,
            folder_id=folder_id,
            name=name,
            description=description,
            preconditions=preconditions,
            priority=priority,
            status=status,
            case_type=case_type,
            owner=owner,
            tags=tags,
            issues=issues,
            automation_status=automation_status,
            custom_fields=custom_fields,
            template=template,
            test_case_steps=test_case_steps,
            feature=feature,
            scenario=scenario,
            background=background,
            module=module,
            keyword=keyword,
            tp_code=tp_code,
            risk_level=risk_level,
            case_kind=case_kind,
        )
    except Exception as e:
        return {"success": False, "error": str(e), "message": f"创建测试用例失败: {str(e)}"}


async def _update_test_case_impl(
    project_identifier: str,
    test_case_identifier: str,
    name: Optional[str] = None,
    description: Optional[str] = None,
    preconditions: Optional[str] = None,
    priority: Optional[str] = None,
    status: Optional[str] = None,
    case_type: Optional[str] = None,
    folder_id: Optional[str] = None,
    owner: Optional[str] = None,
    tags: Optional[list[str]] = None,
    issues: Optional[list[str]] = None,
    automation_status: Optional[str] = None,
    custom_fields: Optional[dict[str, Any]] = None,
    test_case_steps: Optional[list[dict[str, str]]] = None,
    feature: Optional[str] = None,
    scenario: Optional[str] = None,
    background: Optional[str] = None,
    module: Optional[str] = None,
    keyword: Optional[str] = None,
    tp_code: Optional[str] = None,
    risk_level: Optional[str] = None,
    case_kind: Optional[str] = None,
) -> dict[str, Any]:
    """更新测试用例的内部实现"""
    request_data: dict[str, Any] = {}

    if name is not None:
        request_data["name"] = name
    if description is not None:
        request_data["description"] = description
    if preconditions is not None:
        request_data["preconditions"] = preconditions
    if priority is not None:
        request_data["priority"] = priority
    if status is not None:
        request_data["status"] = status
    if case_type is not None:
        request_data["case_type"] = case_type
    if folder_id is not None:
        request_data["folder_id"] = folder_id
    if owner is not None:
        request_data["owner"] = owner
    if tags is not None:
        request_data["tags"] = tags
    if issues is not None:
        request_data["issues"] = issues
    if automation_status is not None:
        request_data["automation_status"] = automation_status
    if custom_fields is not None:
        request_data["custom_fields"] = custom_fields
    if test_case_steps is not None:
        request_data["test_case_steps"] = test_case_steps
    if feature is not None:
        request_data["feature"] = feature
    if scenario is not None:
        request_data["scenario"] = scenario
    if background is not None:
        request_data["background"] = background
    if module is not None:
        request_data["module"] = module
    if keyword is not None:
        request_data["keyword"] = keyword
    if tp_code is not None:
        request_data["tp_code"] = tp_code
    if risk_level is not None:
        request_data["risk_level"] = risk_level
    if case_kind is not None:
        request_data["case_kind"] = case_kind

    if not request_data:
        return {
            "success": False,
            "error": "没有提供任何需要更新的字段",
            "message": "更新测试用例失败：没有提供任何需要更新的字段"
        }

    url = _get_api_url(f"/projects/{project_identifier}/test-cases/{test_case_identifier}")
    response_data = await _make_http_request(method="PATCH", url=url, json_data=request_data)
# pylint: disable  Mi80OmFIVnBZMlhscm9ua3VMazZiWGREVlE9PTphNTIyZWE4MA==

    if response_data.get("success"):
        test_case_data = response_data.get("data", {})
        return {
            "success": True,
            "data": test_case_data,
            "message": f"测试用例 {test_case_data.get('identifier', '')} 更新成功"
        }
    else:
        return {"success": False, "error": "API 返回失败", "message": "更新测试用例失败"}


@tool
async def update_test_case_tool(
    project_identifier: str,
    test_case_identifier: str,
    name: Optional[str] = None,
    description: Optional[str] = None,
    preconditions: Optional[str] = None,
    priority: Optional[str] = None,
    status: Optional[str] = None,
    case_type: Optional[str] = None,
    folder_id: Optional[str] = None,
    owner: Optional[str] = None,
    tags: Optional[list[str]] = None,
    issues: Optional[list[str]] = None,
    automation_status: Optional[str] = None,
    custom_fields: Optional[dict[str, Any]] = None,
    test_case_steps: Optional[list[dict[str, str]]] = None,
    feature: Optional[str] = None,
    scenario: Optional[str] = None,
    background: Optional[str] = None,
    module: Optional[str] = None,
    keyword: Optional[str] = None,
    tp_code: Optional[str] = None,
    risk_level: Optional[str] = None,
    case_kind: Optional[str] = None,
) -> dict[str, Any]:
    """
    更新测试用例工具（通过 HTTP 接口调用）。

    所有字段都是可选的，只更新提供的字段。

    Args:
        project_identifier: 项目标识符，如 'PROJ-001'
        test_case_identifier: 测试用例标识符，如 'TC-1234'
        name: 测试用例名称
        description: 测试用例描述
        preconditions: 前置条件
        priority: 优先级（critical, high, medium, low）
        status: 状态
        case_type: 测试类型
        folder_id: 所属文件夹 UUID（用于移动测试用例）
        owner: 负责人邮箱
        tags: 标签列表
        issues: 关联的 Jira issues
        automation_status: 自动化状态
        custom_fields: 自定义字段
        test_case_steps: 测试步骤列表
        feature: BDD Feature 描述
        scenario: BDD Scenario 描述
        background: BDD Background 描述
        module: 所属模块
        keyword: 关键词（正向, 反向, 边界）
        tp_code: 测试点编号
        risk_level: 风险等级（high, medium, low）
        case_kind: 用例种类（sit, uat, smoke, api）

    Returns:
        dict: 包含更新结果的字典
    """
    try:
        return await _update_test_case_impl(
            project_identifier=project_identifier,
            test_case_identifier=test_case_identifier,
            name=name,
            description=description,
            preconditions=preconditions,
            priority=priority,
            status=status,
            case_type=case_type,
            folder_id=folder_id,
            owner=owner,
            tags=tags,
            issues=issues,
            automation_status=automation_status,
            custom_fields=custom_fields,
            test_case_steps=test_case_steps,
            feature=feature,
            scenario=scenario,
            background=background,
            module=module,
            keyword=keyword,
            tp_code=tp_code,
            risk_level=risk_level,
            case_kind=case_kind,
        )
    except Exception as e:
        return {"success": False, "error": str(e), "message": f"更新测试用例失败: {str(e)}"}


async def _batch_create_test_cases_impl(
    project_identifier: str,
    folder_id: str,
    test_cases: list[dict[str, Any]],
) -> dict[str, Any]:
    """批量创建测试用例的内部实现"""
    if not test_cases:
        return {
            "success": False,
            "error": "测试用例列表为空",
            "message": "批量创建失败：测试用例列表为空"
        }

    results = []
    succeeded = 0
    failed = 0

    for index, test_case_data in enumerate(test_cases):
        try:
            name = test_case_data.get("name")
            if not name:
                results.append({
                    "index": index,
                    "success": False,
                    "error": "测试用例名称不能为空",
                    "data": test_case_data
                })
                failed += 1
                continue

            result = await _create_test_case_impl(
                project_identifier=project_identifier,
                folder_id=folder_id,
                name=name,
                description=test_case_data.get("description"),
                preconditions=test_case_data.get("preconditions"),
                priority=test_case_data.get("priority", "medium"),
                status=test_case_data.get("status", "new"),
                case_type=test_case_data.get("case_type", "functional"),
                owner=test_case_data.get("owner"),
                tags=test_case_data.get("tags"),
                issues=test_case_data.get("issues"),
                automation_status=test_case_data.get("automation_status", "not_automated"),
                custom_fields=test_case_data.get("custom_fields"),
                template=test_case_data.get("template", "test_case"),
                test_case_steps=test_case_data.get("test_case_steps"),
                feature=test_case_data.get("feature"),
                scenario=test_case_data.get("scenario"),
                background=test_case_data.get("background"),
                module=test_case_data.get("module"),
                keyword=test_case_data.get("keyword"),
                tp_code=test_case_data.get("tp_code"),
                risk_level=test_case_data.get("risk_level"),
                case_kind=test_case_data.get("case_kind", "sit"),
            )
# fmt: off  My80OmFIVnBZMlhscm9ua3VMazZiWGREVlE9PTphNTIyZWE4MA==

            results.append({
                "index": index,
                "success": result.get("success", False),
                "data": result.get("data"),
                "error": result.get("error"),
                "message": result.get("message")
            })

            if result.get("success"):
                succeeded += 1
            else:
                failed += 1

        except Exception as e:
            results.append({
                "index": index,
                "success": False,
                "error": str(e),
                "data": test_case_data
            })
            failed += 1

    return {
        "success": True,
        "data": {
            "total": len(test_cases),
            "succeeded": succeeded,
            "failed": failed,
            "results": results
        },
        "message": f"批量创建完成：成功 {succeeded} 个，失败 {failed} 个"
    }


@tool
async def batch_create_test_cases_tool(
    project_identifier: str,
    folder_id: str,
    test_cases: list[dict[str, Any]],
) -> dict[str, Any]:
    """
    批量创建测试用例工具（通过 HTTP 接口调用）。

    每个测试用例的参数与 create_test_case_tool 相同。

    Args:
        project_identifier: 项目标识符，如 'PROJ-001'
        folder_id: 文件夹 UUID
        test_cases: 测试用例列表，每个元素是一个包含测试用例信息的字典

    Returns:
        dict: 包含批量创建结果的字典
    """
    try:
        return await _batch_create_test_cases_impl(
            project_identifier=project_identifier,
            folder_id=folder_id,
            test_cases=test_cases,
        )
    except Exception as e:
        return {
            "success": False,
            "error": str(e),
            "message": f"批量创建测试用例失败: {str(e)}"
        }


# ============================================================================
# 测试用例暂存与导出工具
# ============================================================================

import json as _json
import os
from collections import Counter
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from langchain.tools import tool as _tool

# 延迟导入 staging 避免循环导入
_staging = None

def _get_staging():
    global _staging
    if _staging is None:
        from app.agents.testcase.staging import add, list_ids, list_all, clear, add_uat, list_uat_codes, list_all_uat, clear_uat
        _staging = SimpleNamespace(
            add=add, list_ids=list_ids, list_all=list_all, clear=clear,
            add_uat=add_uat, list_uat_codes=list_uat_codes, list_all_uat=list_all_uat, clear_uat=clear_uat
        )
    return _staging

_HEADER_FILL = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
_HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
_ALIGNMENT_WRAP = Alignment(vertical="top", wrap_text=True)
_ALIGNMENT_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)

_DEFAULT_COLUMN_WIDTHS = {
    "A": 18,
    "B": 35,
    "C": 14,
    "D": 12,
    "E": 10,
    "F": 10,
    "G": 30,
    "H": 40,
    "I": 30,
    "J": 40,
    "K": 20,
}

_EXPORTS_DIR = Path(__file__).resolve().parent.parent.parent.parent.parent / "workspace" / "uploads"
_EXPORTS_DIR.mkdir(parents=True, exist_ok=True)


def _coerce_from_json(value: Any) -> Any:
    """若 value 是看起来像 JSON 的字符串，尝试解析为 list/dict；否则原样返回。"""
    if isinstance(value, str):
        s = value.strip()
        if s.startswith(("[", "{")):
            try:
                return _json.loads(s)
            except Exception:
                return value
    return value


def _flatten_steps(steps: list[dict[str, Any]] | str | None) -> str:
    if not steps:
        return ""
    steps = _coerce_from_json(steps)
    if isinstance(steps, str):
        return steps
    lines: list[str] = []
    for step in steps:
        if isinstance(step, str):
            lines.append(step)
            continue
        if not isinstance(step, dict):
            lines.append(str(step))
            continue
        seq = step.get("seq", step.get("step", len(lines) + 1))
        action = step.get("action", step.get("操作描述", ""))
        target = step.get("target", step.get("操作对象", ""))
        data = step.get("data", "")
        line = f"{seq}. {action}"
        if target:
            line += f" [{target}]"
        if data:
            line += f"（数据：{data}）"
        lines.append(line)
    return "\n".join(lines)


def _flatten_test_data(test_data: dict[str, Any] | str | None) -> str:
    if not test_data:
        return ""
    test_data = _coerce_from_json(test_data)
    if isinstance(test_data, str):
        return test_data
    if isinstance(test_data, list):
        return "\n".join([f"{idx}. {item}" for idx, item in enumerate(test_data, start=1)])
    lines = [f"{k}: {v}" for k, v in test_data.items()]
    return "\n".join(lines)


def _flatten_expected_results(expected_results: list[str] | str | None) -> str:
    if not expected_results:
        return ""
    expected_results = _coerce_from_json(expected_results)
    if isinstance(expected_results, str):
        return expected_results
    return "\n".join([f"{idx}. {r}" for idx, r in enumerate(expected_results, start=1)])


def _flatten_preconditions(preconditions: list[str] | str | None) -> str:
    if not preconditions:
        return ""
    preconditions = _coerce_from_json(preconditions)
    if isinstance(preconditions, str):
        return preconditions
    return "\n".join([f"{idx}. {c}" for idx, c in enumerate(preconditions, start=1)])


def _extract_field(case: dict[str, Any], *keys: str, default: Any = "") -> Any:
    for key in keys:
        if key in case:
            return case[key]
    return default


def _export_test_cases_to_excel_impl(
    test_cases: list[dict[str, Any]],
    output_path: str | Path,
    sheet_name: str = "测试用例",
) -> str:
    """将测试用例列表导出为 Excel 文件（内部实现）。"""
    if not test_cases:
        raise ValueError("测试用例列表为空，无法导出 Excel。")

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    if ws is None:
        raise RuntimeError("无法创建工作表。")
    ws.title = sheet_name

    headers = [
        "用例编号",
        "用例标题",
        "所属模块",
        "用例类型",
        "关键词",
        "优先级",
        "前置条件",
        "测试步骤",
        "测试数据",
        "预期结果",
        "备注",
    ]
    ws.append(headers)

    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _ALIGNMENT_CENTER
        cell.border = _BORDER

    for case in test_cases:
        row = [
            _extract_field(case, "id", "用例编号"),
            _extract_field(case, "title", "用例标题"),
            _extract_field(case, "module", "所属模块"),
            _extract_field(case, "type", "用例类型"),
            _extract_field(case, "keyword", "关键词"),
            _extract_field(case, "priority", "优先级"),
            _flatten_preconditions(_extract_field(case, "preconditions", "前置条件", default=None)),
            _flatten_steps(_extract_field(case, "steps", "测试步骤", default=None)),
            _flatten_test_data(_extract_field(case, "test_data", "测试数据", default=None)),
            _flatten_expected_results(_extract_field(case, "expected_results", "预期结果", default=None)),
            _extract_field(case, "remarks", "备注"),
        ]
        ws.append(row)
        row_idx = ws.max_row
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.alignment = _ALIGNMENT_WRAP
            cell.border = _BORDER

    for col_letter, width in _DEFAULT_COLUMN_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width

    ws.row_dimensions[1].height = 24
    for row_idx in range(2, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 60

    wb.save(str(output_path))
    return str(output_path.resolve())


@_tool
async def stage_testcases(test_cases: list, batch_note: str = "") -> str:
    """
    将本轮生成的一批测试用例追加到"会话暂存区"。每当你生成完一个模块或一批用例，
    必须立即调用本工具入库，而不是等到导出时才一次性传入。

    去重规则：按 id / 用例编号 去重，后写覆盖先写（允许你修正某条后重新暂存）。
    该暂存区仅在当前进程内存中维护，跨服务重启会清空。

    Args:
        test_cases: 本批次测试用例列表，每条为 dict，字段与 export_testcases_to_excel 一致
            （id, title, module, type, keyword, priority, preconditions, steps, test_data,
             expected_results, remarks）。其中 keyword 枚举值为：正向 / 反向 / 边界。
        batch_note: 可选备注，用于你自己标记这批次的主题，例如"SDK集成模块"、"Android专项"。

    Returns:
        JSON 字符串，包含 staged_count（本次入库条数）、total_count（累计总数）、
        total_ids（累计所有用例编号的清单）。你应把 total_count 告知用户。
    """
    if not isinstance(test_cases, list):
        return _json.dumps({"error": "test_cases 必须是 list[dict]"}, ensure_ascii=False)

    # 入库前实时质量拦截：模糊词检查
    fuzzy_words = ["正确", "成功", "正常", "验证成功", "功能正常", "页面正确"]
    rejected = []
    accepted = []
    for case in test_cases:
        expected = str(case.get("expected_results", ""))
        found_fuzzy = None
        for word in fuzzy_words:
            if word in expected:
                found_fuzzy = word
                break
        if found_fuzzy:
            rejected.append({
                "id": case.get("id", "unknown"),
                "title": case.get("title", "")[:40],
                "fuzzy_word": found_fuzzy
            })
        else:
            accepted.append(case)

    staged_count = len(accepted)
    total = _get_staging().add(accepted)
    ids = _get_staging().list_ids()

    result = {
        "staged_count": staged_count,
        "total_count": total,
        "total_ids": ids,
        "batch_note": batch_note,
        "message": f"本批次入库 {staged_count} 条，暂存区累计 {total} 条。",
    }
    if rejected:
        result["rejected_count"] = len(rejected)
        result["rejected_cases"] = rejected
        result["warning"] = f"有 {len(rejected)} 条用例因预期结果含模糊词被拒绝入库，请重写后重新提交。"

    return _json.dumps(result, ensure_ascii=False)


@_tool
async def list_staged_testcases() -> str:
    """
    查询会话暂存区中当前累计的所有测试用例编号清单与总数。

    **强制使用场景**：当用户要求导出 Excel 前，必须先调用本工具向用户展示累计清单，
    经用户确认数量正确后，才能调用 export_all_testcases 触发导出。

    Returns:
        JSON 字符串，包含 total_count 与 total_ids。
    """
    ids = _get_staging().list_ids()
    return _json.dumps(
        {"total_count": len(ids), "total_ids": ids},
        ensure_ascii=False,
    )


@_tool
async def export_all_testcases(requirement_name: str = "", sheet_name: str = "测试用例", clear_after: bool = False, case_kind: str = "sit") -> str:
    """
    **一键导出全部**：将会话暂存区中累计的所有测试用例一次性导出为 Excel。
    本工具不接受 test_cases 参数，确保不会因你漏传而丢条——**只要用户要求导出
    测试用例 Excel，就应该使用本工具而不是 export_testcases_to_excel**。

    文件命名规则：
      - SIT 模式（默认）：{requirement_name}_SIT测试用例_{时间戳}.xlsx
      - 冒烟模式：{requirement_name}_冒烟测试_{时间戳}.xlsx

    Args:
        requirement_name: 需求名称，用于文件命名（**必须从用户消息或对话上下文中提取**，
            禁止留空或传"测试用例"）。
        sheet_name: 工作表名称，默认 "测试用例"。
        clear_after: 导出成功后是否清空暂存区。默认 False（保留，允许用户再次导出或
            继续追加）。当用户明确表示"这一轮结束、开始新需求"时，传 True。
        case_kind: 用例种类，'sit'（默认）或 'smoke'。当前为冒烟测试模式时，
            **必须**传 'smoke'，文件名会生成 `_冒烟测试_` 后缀。

    Returns:
        JSON 字符串，包含 filename、exported_count、message，以及可选 file_id。
    """
    cases = _get_staging().list_all()
    if not cases:
        return _json.dumps(
            {"error": "暂存区为空。请先使用 stage_testcases 工具把本轮生成的用例入库，再调用本工具。"},
            ensure_ascii=False,
        )

    # 导出前质量门禁检查
    quality_errors = []
    total = len(cases)
    if total == 0:
        return _json.dumps({"error": "暂存区为空"}, ensure_ascii=False)

    # 允许通过环境变量临时绕过质量门禁（用于紧急情况导出已生成数据）
    _SKIP_QUALITY_GATE = os.getenv("SKIP_QUALITY_GATE", "0") == "1"

    # 1. 关键词比例检查
    zhengxiang = sum(1 for c in cases if c.get("keyword") == "正向")
    fanxiang = sum(1 for c in cases if c.get("keyword") == "反向")
    bianjie = sum(1 for c in cases if c.get("keyword") == "边界")

    if zhengxiang / total > 0.45:
        quality_errors.append(f"正向用例占比 {zhengxiang/total*100:.1f}% > 45%，必须削减正向或补充反向/边界用例")
    if fanxiang / total < 0.30:
        quality_errors.append(f"反向用例占比 {fanxiang/total*100:.1f}% < 30%，必须补充反向用例")
    if bianjie / total < 0.10:
        quality_errors.append(f"边界用例占比 {bianjie/total*100:.1f}% < 10%，必须补充边界用例")

    # 2. 模糊词检查（阈值放宽到10%，095910版本经验值）
    fuzzy_words = ["正确", "成功", "正常", "验证成功", "功能正常", "页面正确"]
    fuzzy_cases = []
    for c in cases:
        expected = str(c.get("expected_results", ""))
        for word in fuzzy_words:
            if word in expected:
                fuzzy_cases.append(c.get("id", "unknown"))
                break

    if len(fuzzy_cases) / total > 0.10:
        quality_errors.append(f"模糊预期结果占比 {len(fuzzy_cases)/total*100:.1f}% > 10%，用例 {fuzzy_cases[:5]} 等必须重写。标准要求：模糊词占比 ≤ 10%")

    # 3. 9维覆盖检查 - 检查是否有DB/并发/安全等标记（"有即可"原则，不强制比例）
    has_db = any("DB" in str(c.get("remarks", "")) or "数据库" in str(c.get("remarks", "")) for c in cases)
    has_concurrent = any("并发" in str(c.get("remarks", "")) or "幂等" in str(c.get("remarks", "")) for c in cases)
    has_security = any("安全" in str(c.get("remarks", "")) or "SQL" in str(c.get("remarks", "")) for c in cases)

    if not has_db:
        quality_errors.append("缺少DB异常类用例，涉及写操作的模块必须补充")
    if not has_concurrent:
        quality_errors.append("缺少并发安全类用例，核心写入模块必须补充")
    if not has_security:
        quality_errors.append("缺少字段级安全类用例，涉及用户输入的模块必须补充")

    # 4. 字段级安全用例数量检查（回退到"有即可"：≥1条，不强制比例）
    security_count = sum(1 for c in cases if "安全" in str(c.get("remarks", "")) or "SQL" in str(c.get("remarks", "")) or "XSS" in str(c.get("remarks", "")))
    if security_count < 1:
        quality_errors.append("字段级安全用例至少需有1条，涉及用户输入的模块必须补充")

    # 4.5 并发安全用例数量检查（回退到"有即可"：≥1条，不强制比例）
    concurrent_count = sum(1 for c in cases if "并发" in str(c.get("remarks", "")) or "幂等" in str(c.get("remarks", "")))
    if concurrent_count < 1:
        quality_errors.append("并发安全用例至少需有1条，核心写入模块必须补充")

    # 5. 模块分布均衡性检查（放宽到≤25%，095910版本经验值）
    modules = Counter(str(c.get("module", "")) for c in cases if c.get("module"))
    if modules:
        max_module_count = modules.most_common(1)[0][1]
        min_module_count = modules.most_common()[-1][1]
        if max_module_count / total > 0.25:
            max_module = modules.most_common(1)[0][0]
            quality_errors.append(f"模块分布不均衡：'{max_module}' 模块有 {max_module_count} 条用例（占比 {max_module_count/total*100:.1f}% > 25% 标准），请削减该模块正向用例或补充其他模块覆盖")
        elif min_module_count > 0 and max_module_count / min_module_count > 5:
            max_module = modules.most_common(1)[0][0]
            min_module = modules.most_common()[-1][0]
            quality_errors.append(f"模块间差异过大：'{max_module}'（{max_module_count}条）与'{min_module}'（{min_module_count}条）差异 {max_module_count/min_module_count:.1f} 倍 > 5倍，请均衡覆盖")

    if quality_errors:
        import logging
        logger = logging.getLogger(__name__)
        if _SKIP_QUALITY_GATE:
            logger.warning("[export_all_testcases] 质量门禁有 %d 项警告，但 SKIP_QUALITY_GATE=1 已绕过。警告列表：%s", len(quality_errors), quality_errors)
        else:
            return _json.dumps({
                "error": "质量门禁检查未通过，禁止导出。请返回补充以下问题：",
                "quality_errors": quality_errors,
                "total_cases": total,
                "zhengxiang": zhengxiang,
                "fanxiang": fanxiang,
                "bianjie": bianjie,
            }, ensure_ascii=False)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe_name = requirement_name.strip() if requirement_name else "测试用例"
    suffix = "冒烟测试" if case_kind == "smoke" else "SIT测试用例"
    filename = f"{safe_name}_{suffix}_{timestamp}.xlsx"
    output_path = _EXPORTS_DIR / filename

    _export_test_cases_to_excel_impl(cases, output_path, sheet_name)

    exported_count = len(cases)
    if clear_after:
        _get_staging().clear()

    result = {
        "filename": filename,
        "exported_count": exported_count,
        "message": f"已一次性导出 {exported_count} 条用例到 {filename}，可通过前端导出按钮下载。",
    }
    if clear_after:
        result["staging_cleared"] = True
    return _json.dumps(result, ensure_ascii=False)


@_tool
async def clear_staged_testcases() -> str:
    """
    清空会话暂存区。仅在用户明确表示"开始新一轮需求、放弃之前的用例"时调用。
    Returns:
        JSON 字符串，包含被清除的条数 cleared_count。
    """
    n = _get_staging().clear()
    return _json.dumps({"cleared_count": n, "message": f"已清空暂存区（{n} 条）。"}, ensure_ascii=False)


# ============================================================================
# UAT 业务场景工具（独立暂存通道，不与 SIT 共用）
# ============================================================================

_UAT_HEADER_BIZ = ["业务编号", "分类", "业务类型", "业务说明"]
_UAT_HEADER_STEP = [
    "序号",
    "前置条件",
    "场景说明",
    "流程节点（操作步骤）",
    "涉及中心(第三方)",
    "预期结果",
    "测试结果",
    "测试人员",
]
_UAT_COLUMN_WIDTHS = {
    "A": 10,
    "B": 28,
    "C": 24,
    "D": 40,
    "E": 18,
    "F": 40,
    "G": 12,
    "H": 12,
}


def _uat_step_field(step: dict[str, Any], *keys: str, default: Any = "") -> Any:
    for k in keys:
        if k in step and step[k] not in (None, ""):
            return step[k]
    return default


def _export_uat_scenarios_to_excel_impl(
    scenarios: list[dict[str, Any]],
    output_path: str | Path,
    sheet_name: str = "UAT验收用例",
) -> str:
    """按 UAT 业务场景样例.xlsx 的两段式结构导出 Excel（内部实现）。"""
    if not scenarios:
        raise ValueError("UAT 业务场景列表为空，无法导出 Excel。")

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    if ws is None:
        raise RuntimeError("无法创建工作表。")
    ws.title = sheet_name

    for col_letter, width in _UAT_COLUMN_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width

    def _write_header_row(values: list[str], col_count: int) -> int:
        ws.append(values + [None] * (col_count - len(values)))
        row_idx = ws.max_row
        for col_idx in range(1, col_count + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if col_idx <= len(values):
                cell.fill = _HEADER_FILL
                cell.font = _HEADER_FONT
                cell.alignment = _ALIGNMENT_CENTER
                cell.border = _BORDER
        return row_idx

    def _write_data_row(values: list[Any], col_count: int) -> int:
        ws.append(values + [None] * (col_count - len(values)))
        row_idx = ws.max_row
        for col_idx in range(1, col_count + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.alignment = _ALIGNMENT_WRAP
            cell.border = _BORDER
        return row_idx

    for sc in scenarios:
        biz_code = _extract_field(sc, "biz_code", "业务编号")
        category = _extract_field(sc, "category", "分类")
        biz_type = _extract_field(sc, "biz_type", "业务类型")
        biz_desc = _extract_field(sc, "biz_description", "业务说明")
        scenario_desc = _extract_field(sc, "scenario_description", "业务场景说明")

        steps = sc.get("steps") or sc.get("流程节点") or []
        steps = _coerce_from_json(steps)
        if not isinstance(steps, list):
            steps = []

        # 段 1：业务头
        _write_header_row(_UAT_HEADER_BIZ, col_count=8)
        _write_data_row([biz_code, category, biz_type, biz_desc], col_count=8)

        # "业务场景说明" 单独一行（首列写标签，第二列起合并写内容）
        ws.append(["业务场景说明", scenario_desc, None, None, None, None, None, None])
        meta_row = ws.max_row
        label_cell = ws.cell(row=meta_row, column=1)
        label_cell.fill = _HEADER_FILL
        label_cell.font = _HEADER_FONT
        label_cell.alignment = _ALIGNMENT_CENTER
        label_cell.border = _BORDER
        try:
            ws.merge_cells(start_row=meta_row, start_column=2, end_row=meta_row, end_column=8)
        except Exception:
            pass
        for col_idx in range(2, 9):
            c = ws.cell(row=meta_row, column=col_idx)
            c.alignment = _ALIGNMENT_WRAP
            c.border = _BORDER

        # 空行
        ws.append([None] * 8)

        # 段 2：流程节点表
        _write_header_row(_UAT_HEADER_STEP, col_count=8)
        for idx, step in enumerate(steps, start=1):
            if not isinstance(step, dict):
                step = {"action": str(step)}
            row = [
                _uat_step_field(step, "seq", "序号", default=idx),
                _uat_step_field(step, "preconditions", "前置条件"),
                _uat_step_field(step, "scenario_note", "场景说明"),
                _uat_step_field(step, "action", "流程节点", "操作步骤"),
                _uat_step_field(step, "involved_systems", "涉及中心", "涉及中心(第三方)", "涉及中心（第三方）"),
                _uat_step_field(step, "expected_result", "预期结果"),
                "",  # 测试结果（执行阶段填）
                "",  # 测试人员（执行阶段填）
            ]
            _write_data_row(row, col_count=8)

        # 场景间空行
        ws.append([None] * 8)

    # 行高：表头 24，其它默认让 wrap_text 自动撑开
    for r in range(1, ws.max_row + 1):
        if ws.row_dimensions[r].height is None:
            ws.row_dimensions[r].height = None

    wb.save(str(output_path))
    return str(output_path.resolve())


@_tool
async def stage_uat_scenarios(scenarios: list, batch_note: str = "") -> str:
    """
    将本轮生成的一批 **UAT 业务场景** 追加到 UAT 暂存区。每生成完一组业务场景，
    立即调用本工具入库；不要等到导出时才一次性传入。SIT 用例请改用 stage_testcases。

    去重规则：按 biz_code（业务编号，如 X001/X002）去重，后写覆盖先写。

    Args:
        scenarios: UAT 业务场景列表，每条为 dict，字段如下：
            - biz_code               业务编号（必填，格式 X001、X002 …，全局唯一流水号）
            - category               分类（必填，如"业务单据——盘盈盘亏单"）
            - biz_type               业务类型（必填，如"盘盈盘亏"）
            - biz_description        业务说明（必填，一句话说清这个业务做什么）
            - scenario_description   业务场景说明（必填，描述具体场景，如"盘盈盘亏单，盘盈盘亏（盘盈）"）
            - steps                  流程节点列表（**必填，原生 list，禁止序列化为 JSON 字符串**），
                                     每个节点为 dict：
                * seq               序号（数字，从 1 开始）
                * preconditions     前置条件（可空；同场景内首节点写出后，后续节点可留空）
                * scenario_note     场景说明（可空；首节点写出后，后续节点可留空）
                * action            流程节点 / 操作步骤（必填，可含换行 \n 与中文标点）
                * involved_systems  涉及中心 / 第三方（可空，如 "OA"、"ERP"）
                * expected_result   预期结果（必填，明确可验证）
        batch_note: 可选备注，标记本批次主题，例如"主流程"、"异常分支"。

    Returns:
        JSON 字符串，含 staged_count（本次入库）、total_count（累计）、
        total_codes（累计业务编号清单）。
    """
    if not isinstance(scenarios, list):
        return _json.dumps({"error": "scenarios 必须是 list[dict]"}, ensure_ascii=False)
    staged_count = len(scenarios)
    total = _get_staging().add_uat(scenarios)
    codes = _get_staging().list_uat_codes()
    return _json.dumps(
        {
            "staged_count": staged_count,
            "total_count": total,
            "total_codes": codes,
            "batch_note": batch_note,
            "message": f"本批次入库 {staged_count} 个 UAT 场景，暂存区累计 {total} 个。",
        },
        ensure_ascii=False,
    )


@_tool
async def list_staged_uat_scenarios() -> str:
    """
    查询 UAT 暂存区当前累计的所有业务场景编号清单与总数。

    **强制使用场景**：用户要求导出 UAT Excel 前，必须先调用本工具向用户展示累计清单，
    经用户确认数量正确后，才能调用 export_all_uat_scenarios 触发导出。
    """
    codes = _get_staging().list_uat_codes()
    return _json.dumps(
        {"total_count": len(codes), "total_codes": codes},
        ensure_ascii=False,
    )


@_tool
async def export_all_uat_scenarios(requirement_name: str = "", sheet_name: str = "UAT验收用例", clear_after: bool = False) -> str:
    """
    **一键导出 UAT 业务场景**：将 UAT 暂存区中累计的所有业务场景一次性导出为 Excel。
    导出格式严格对齐"UAT 业务场景样例.xlsx"模板（业务头 + 业务场景说明 + 流程节点表，
    多场景纵向堆叠）。

    文件命名规则：{requirement_name}_UAT验收用例_{时间戳}.xlsx

    Args:
        requirement_name: 需求名称，**必须从用户消息或对话上下文中提取**，禁止留空或
            传"测试用例"等无意义默认值。例如用户说"帮我做盘盈盘亏单的验收用例"，传入
            "盘盈盘亏单"。
        sheet_name: 工作表名称，默认 "UAT验收用例"。
        clear_after: 导出后是否清空 UAT 暂存区。默认 False。

    Returns:
        JSON 字符串，包含 filename、exported_count、message，以及可选 file_id。
    """
    scenarios = _get_staging().list_all_uat()
    if not scenarios:
        return _json.dumps(
            {"error": "UAT 暂存区为空。请先使用 stage_uat_scenarios 把本轮生成的业务场景入库，再调用本工具。"},
            ensure_ascii=False,
        )

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe_name = requirement_name.strip() if requirement_name else "需求"
    filename = f"{safe_name}_UAT验收用例_{timestamp}.xlsx"
    output_path = _EXPORTS_DIR / filename

    _export_uat_scenarios_to_excel_impl(scenarios, output_path, sheet_name)

    exported_count = len(scenarios)
    if clear_after:
        _get_staging().clear_uat()

    result = {
        "filename": filename,
        "exported_count": exported_count,
        "message": f"已导出 {exported_count} 个 UAT 业务场景到 {filename}，可通过前端导出按钮下载。",
    }
    if clear_after:
        result["staging_cleared"] = True
    return _json.dumps(result, ensure_ascii=False)


@_tool
async def clear_staged_uat_scenarios() -> str:
    """清空 UAT 暂存区。仅在用户明确表示"重做 UAT、放弃之前的场景"时调用。"""
    n = _get_staging().clear_uat()
    return _json.dumps({"cleared_count": n, "message": f"已清空 UAT 暂存区（{n} 个场景）。"}, ensure_ascii=False)
