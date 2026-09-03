"""
测试用例管理 API

提供测试用例的 CRUD 操作接口
参考: https://www.browserstack.com/docs/test-management/api-reference/test-cases
"""

import logging
from datetime import datetime
from typing import Optional, Union, Any, List
from uuid import UUID
import os
from pathlib import Path

from fastapi import APIRouter, Query, status, Request, Response, Body
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession

from app.api.deps import (
    TestCaseServiceDep,
    PaginationDep,
    CurrentUserIdDep,
    DbSessionDep,
    ExportServiceDep,
)
from app.schemas.test_case import (
    TestCaseCreate, TestCaseUpdate, TestCaseInfo, TestCaseMinifiedInfo,
    BulkTestCaseRequest, BulkEditWithOperationsRequest, BulkDeleteRequest,
    BulkOperationResponse, ExportBDDRequest, ExportBDDResponse,
    ExportStatusResponse, TestCaseHistoryResponse
)
from app.schemas.common import SuccessResponse, MessageResponse
from app.schemas.pagination import PaginatedResponse, PaginationInfo
from app.schemas.enums import Priority, TestCaseState, TestCaseType
from app.config.settings import settings
from app.utils.exceptions import BadRequestException
from app.utils.http_headers import build_content_disposition

router = APIRouter(prefix="/projects/{project_identifier}")

# ============ 测试用例列表和查询接口 ============

@router.get(
    "/test-cases",
    response_model=PaginatedResponse[Union[TestCaseInfo, TestCaseMinifiedInfo]],
    summary="获取测试用例列表",
    description="获取项目下的所有测试用例列表，支持过滤和分页",
)
async def get_test_cases(
    project_identifier: str,
    request: Request,
    service: TestCaseServiceDep,
    pagination: PaginationDep,
    # 精简模式
    minify: bool = Query(False, description="是否返回精简数据"),
    # 根据 ID 过滤
    id: Optional[str] = Query(None, description="测试用例标识符列表，逗号分隔，如 TC-1234,TC-1235"),
    # 属性过滤
    folder_id: Optional[str] = Query(None, description="文件夹 ID 列表，逗号分隔"),
    status: Optional[str] = Query(None, description="状态列表，逗号分隔，如 active,draft"),
    priority: Optional[str] = Query(None, description="优先级列表，逗号分隔，如 high,medium"),
    case_type: Optional[str] = Query(None, description="测试类型列表，逗号分隔，如 functional,regression"),
    case_kind: Optional[str] = Query(None, description="用例种类列表，逗号分隔，如 sit,uat,smoke,api"),
    owner: Optional[str] = Query(None, description="负责人邮箱列表，逗号分隔"),
    tags: Optional[str] = Query(None, description="标签列表，逗号分隔"),
    # 时间过滤
    updated_after: Optional[datetime] = Query(None, description="更新时间晚于"),
    updated_before: Optional[datetime] = Query(None, description="更新时间早于"),
    # Jira 集成
    issue_ids: Optional[str] = Query(None, description="关联的 Jira issue ID 列表，逗号分隔"),
    issue_type: Optional[str] = Query(None, description="issue 类型，如 jira"),
    # 搜索
    search: Optional[str] = Query(None, description="搜索关键词，匹配测试用例名称、标识符、描述"),
):
    """
    获取测试用例列表

    支持多种过滤条件：
    - **minify**: 是否返回精简数据（仅包含基本字段）
    - **id**: 根据测试用例标识符过滤
    - **folder_id**: 根据文件夹过滤
    - **status**: 根据状态过滤
    - **priority**: 根据优先级过滤
    - **case_type**: 根据测试类型过滤
    - **owner**: 根据负责人过滤
    - **tags**: 根据标签过滤
    - **issue_ids**: 根据关联的 Jira issue 过滤

    多个值用逗号分隔，同一参数内的多个值为 OR 关系，不同参数之间为 AND 关系

    自定义字段过滤：使用 custom_fields[字段名]=值1,值2 格式
    """
    # 解析自定义字段过滤
    custom_fields = {}
    for key, value in request.query_params.items():
        if key.startswith("custom_fields[") and key.endswith("]"):
            field_name = key[14:-1]  # 提取字段名
            custom_fields[field_name] = value.split(",") if value else []

    # 构建过滤参数
    filters = {
        "test_case_ids": id.split(",") if id else None,
        "folder_ids": folder_id.split(",") if folder_id else None,
        "statuses": status.split(",") if status else None,
        "priorities": priority.split(",") if priority else None,
        "case_types": case_type.split(",") if case_type else None,
        "case_kinds": case_kind.split(",") if case_kind else None,
        "owners": owner.split(",") if owner else None,
        "tags": tags.split(",") if tags else None,
        "issue_ids": issue_ids.split(",") if issue_ids else None,
        "issue_type": issue_type,
        "custom_fields": custom_fields if custom_fields else None,
        "updated_after": updated_after,
        "updated_before": updated_before,
        "search": search,
    }

    offset = (pagination.page - 1) * pagination.page_size
    test_cases, total = await service.get_test_cases(
        project_identifier,
        offset,
        pagination.page_size,
        minify=minify,
        **filters
    )

    total_pages = (total + pagination.page_size - 1) // pagination.page_size if total > 0 else 1
    base_url = f"{settings.api_prefix}/projects/{project_identifier}/test-cases"
# pragma: no cover  MC80OmFIVnBZMlhscm9ua3VMazZhM2xhVUE9PTo4NTMxNTdhNQ==

    prev_url = None
    if pagination.page > 1:
        prev_url = f"{base_url}?p={pagination.page - 1}&page_size={pagination.page_size}"

    next_url = None
    if pagination.page < total_pages:
        next_url = f"{base_url}?p={pagination.page + 1}&page_size={pagination.page_size}"

    return PaginatedResponse(
        success=True,
        data=test_cases,
        info=PaginationInfo(
            page=pagination.page,
            page_size=pagination.page_size,
            count=len(test_cases),
            total=total,
            prev=prev_url,
            next=next_url,
        ),
    )

# ============ 测试用例详情接口 ============

@router.get(
    "/folders/{folder_id}/test-cases",
    response_model=PaginatedResponse[Union[TestCaseInfo, TestCaseMinifiedInfo]],
    summary="获取文件夹下的测试用例列表",
    description="获取指定文件夹下的所有测试用例列表，支持过滤和分页",
)
async def get_folder_test_cases(
    project_identifier: str,
    folder_id: str,
    service: TestCaseServiceDep,
    pagination: PaginationDep,
    # 精简模式
    minify: bool = Query(False, description="是否返回精简数据"),
    # 属性过滤
    status: Optional[str] = Query(None, description="状态列表，逗号分隔，如 active,draft"),
    priority: Optional[str] = Query(None, description="优先级列表，逗号分隔，如 high,medium"),
    case_type: Optional[str] = Query(None, description="测试类型列表，逗号分隔，如 functional,regression"),
    case_kind: Optional[str] = Query(None, description="用例种类列表，逗号分隔，如 sit,uat,smoke,api"),
    owner: Optional[str] = Query(None, description="负责人邮箱列表，逗号分隔"),
    tags: Optional[str] = Query(None, description="标签列表，逗号分隔"),
    # 搜索
    search: Optional[str] = Query(None, description="搜索关键词，匹配测试用例名称、标识符、描述"),
):
    """获取文件夹下的测试用例列表"""
    filters = {
        "folder_ids": [folder_id],
        "statuses": status.split(",") if status else None,
        "priorities": priority.split(",") if priority else None,
        "case_types": case_type.split(",") if case_type else None,
        "case_kinds": case_kind.split(",") if case_kind else None,
        "owners": owner.split(",") if owner else None,
        "tags": tags.split(",") if tags else None,
        "search": search,
    }

    offset = (pagination.page - 1) * pagination.page_size
    test_cases, total = await service.get_test_cases(
        project_identifier,
        offset,
        pagination.page_size,
        minify=minify,
        **filters
    )

    total_pages = (total + pagination.page_size - 1) // pagination.page_size if total > 0 else 1
    base_url = f"{settings.api_prefix}/projects/{project_identifier}/folders/{folder_id}/test-cases"

    prev_url = None
    if pagination.page > 1:
        prev_url = f"{base_url}?p={pagination.page - 1}&page_size={pagination.page_size}"

    next_url = None
    if pagination.page < total_pages:
        next_url = f"{base_url}?p={pagination.page + 1}&page_size={pagination.page_size}"

    return PaginatedResponse(
        success=True,
        data=test_cases,
        info=PaginationInfo(
            page=pagination.page,
            page_size=pagination.page_size,
            count=len(test_cases),
            total=total,
            prev=prev_url,
            next=next_url,
        ),
    )

@router.get(
    "/test-cases/export-excel",
    summary="导出测试用例为 Excel",
    description="导出测试用例为 Excel 文件，支持导出全部、按文件夹或按选中ID导出",
)
async def export_test_cases_excel(
    project_identifier: str,
    request: Request,
    service: TestCaseServiceDep,
    folder_id: Optional[str] = Query(None, description="文件夹 ID，导出该文件夹下的测试用例"),
    test_case_ids: Optional[str] = Query(None, description="测试用例 ID 列表，逗号分隔，导出指定的测试用例"),
    export_type: str = Query("test_cases", description="导出类型: test_cases=测试用例, req_analysis=需求分析报告"),
    requirement_name: Optional[str] = Query(None, description="需求名称，用于文件命名"),
) -> StreamingResponse:
    """
    导出测试用例为 Excel 文件

    - **folder_id**: 指定文件夹 ID，导出该文件夹下的测试用例
    - **test_case_ids**: 指定测试用例 ID 列表，导出指定的测试用例
    - **export_type**: 导出类型，test_cases 导出测试用例，req_analysis 导出需求分析报告

    如果都不指定，则导出项目下的所有测试用例
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
    import io

    # 获取测试用例数据
    # get_test_cases 返回 (list[TestCaseInfo], int) 元组
    offset = 0
    limit = 10000  # 大量导出

    if folder_id:
        test_cases, total = await service.get_test_cases(
            project_identifier,
            offset=offset,
            limit=limit,
            folder_ids=[folder_id],
        )
    elif test_case_ids:
        tc_id_list = [id.strip() for id in test_case_ids.split(",") if id.strip()]
        test_cases, total = await service.get_test_cases(
            project_identifier,
            offset=offset,
            limit=limit,
            test_case_ids=tc_id_list,
        )
    else:
        test_cases, total = await service.get_test_cases(
            project_identifier,
            offset=offset,
            limit=limit,
        )

    if not test_cases:
        raise BadRequestException("没有测试用例可导出")

    # 创建工作簿
    wb = Workbook()
    ws = wb.active
    if ws is None:
        raise RuntimeError("无法创建工作表")

    # 样式定义
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    alignment_wrap = Alignment(vertical="top", wrap_text=True)
    alignment_center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    if export_type == "req_analysis":
        # 导出需求分析报告格式
        ws.title = "需求分析报告"
        headers = [
            "需求编号",
            "需求名称",
            "所属模块",
            "功能描述",
            "测试要点",
            "优先级",
            "风险等级",
            "状态",
            "备注",
        ]
        column_widths = {"A": 15, "B": 30, "C": 18, "D": 40, "E": 35, "F": 10, "G": 10, "H": 10, "I": 20}
    else:
        # 导出测试用例格式
        ws.title = "测试用例"
        headers = [
            "用例编号",
            "用例标题",
            "所属模块",
            "用例种类",  # SIT/UAT/冒烟/API
            "用例类型",  # 功能/冒烟/回归/安全等
            "关键词",
            "优先级",
            "前置条件",
            "测试步骤",
            "测试数据",
            "预期结果",
            "备注",
        ]
        column_widths = {"A": 18, "B": 35, "C": 14, "D": 10, "E": 12, "F": 10, "G": 10, "H": 30, "I": 40, "J": 30, "K": 40, "L": 20}

    # 写入表头
    ws.append(headers)
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = alignment_center
        cell.border = border

    # 写入数据 - test_cases 是 TestCaseInfo 对象列表（Pydantic模型）
    for tc in test_cases:
        # 将 Pydantic 模型转为字典
        if hasattr(tc, 'model_dump'):
            tc_dict = tc.model_dump()
        elif hasattr(tc, 'dict'):
            tc_dict = tc.dict()
        else:
            tc_dict = tc

        if export_type == "req_analysis":
            # 需求分析报告格式
            row = [
                tc_dict.get("identifier", ""),
                tc_dict.get("name", ""),
                tc_dict.get("module", ""),
                _strip_html(tc_dict.get("description", "")),
                _strip_html(tc_dict.get("preconditions", "")),
                tc_dict.get("priority", ""),
                tc_dict.get("risk_level", ""),
                tc_dict.get("status", ""),
                ", ".join(tc_dict.get("tags", [])) if tc_dict.get("tags") else "",
            ]
        else:
            # 测试用例格式
            # 处理测试步骤
            steps = tc_dict.get("test_case_steps", [])
            steps_text = ""
            expected_results = ""
            if steps:
                step_lines = []
                result_lines = []
                for idx, step in enumerate(steps, start=1):
                    step_desc = step.get("step", "") if isinstance(step, dict) else getattr(step, 'step', '')
                    result = step.get("result") if isinstance(step, dict) else getattr(step, 'result', None)
                    result = result or ""  # 处理 None
                    step_lines.append(f"{idx}. {step_desc}")
                    if result and result.strip():
                        result_lines.append(f"{idx}. {result}")
                steps_text = "\n".join(step_lines)
                expected_results = "\n".join(result_lines)

            # 处理测试数据
            test_data = ""
            custom_fields = tc_dict.get("custom_fields", {})
            if custom_fields:
                if isinstance(custom_fields, dict):
                    test_data_lines = [f"{k}: {v}" for k, v in custom_fields.items()]
                    test_data = "\n".join(test_data_lines)
                else:
                    test_data = str(custom_fields)

            row = [
                tc_dict.get("identifier", ""),
                _clean_title(
                    tc_dict.get("name", ""),
                    tc_dict.get("identifier", ""),
                    tc_dict.get("case_type", ""),
                    tc_dict.get("keyword", "")
                ),
                tc_dict.get("module", ""),
                tc_dict.get("case_kind", "sit"),  # 用例种类: SIT/UAT/冒烟/API
                _CASE_TYPE_LABELS.get(tc_dict.get("case_type", ""), tc_dict.get("case_type", "")),  # 用例类型中文
                tc_dict.get("keyword", ""),
                tc_dict.get("priority", ""),
                _strip_html(tc_dict.get("preconditions", "")),
                steps_text,
                test_data,
                expected_results,
                ", ".join(tc_dict.get("tags", [])) if tc_dict.get("tags") else "",
            ]

        ws.append(row)
        row_idx = ws.max_row
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.alignment = alignment_wrap
            cell.border = border

    # 设置列宽
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width

    # 设置行高
    ws.row_dimensions[1].height = 24
    for row_idx in range(2, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 60

    # 保存到内存
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    # 文件名
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    if export_type == "req_analysis":
        filename = f"需求分析报告_{project_identifier}_{timestamp}.xlsx"
    else:
        # 如果有需求名称，则包含在文件名中
        if requirement_name:
            safe_name = requirement_name.strip()
            filename = f"{safe_name}_{project_identifier}_{timestamp}.xlsx"
        else:
            filename = f"测试用例_{project_identifier}_{timestamp}.xlsx"

    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": build_content_disposition(filename)}
    )

@router.get(
    "/test-cases/{test_case_identifier}",
    response_model=SuccessResponse[TestCaseInfo],
    summary="获取测试用例详情",
    description="获取指定测试用例的详细信息",
)
async def get_test_case(
    project_identifier: str,
    test_case_identifier: str,
    service: TestCaseServiceDep,
) -> SuccessResponse[TestCaseInfo]:
    """获取测试用例详情"""
    test_case = await service.get_test_case(project_identifier, test_case_identifier)
    return SuccessResponse(success=True, data=test_case)

# fmt: off  MS80OmFIVnBZMlhscm9ua3VMazZhM2xhVUE9PTo4NTMxNTdhNQ==

# ============ 创建测试用例接口 ============

@router.post(
    "/folders/{folder_id}/test-cases",
    response_model=SuccessResponse[TestCaseInfo],
    status_code=status.HTTP_201_CREATED,
    summary="在文件夹中创建测试用例",
    description="在指定文件夹下创建新的测试用例（支持普通测试用例和 BDD 测试用例）",
)
async def create_test_case_in_folder(
    project_identifier: str,
    folder_id: UUID,
    data: TestCaseCreate,
    service: TestCaseServiceDep,
    current_user_id: CurrentUserIdDep,
    db: DbSessionDep,
) -> SuccessResponse[TestCaseInfo]:
    """
    在文件夹中创建测试用例

    支持两种模板类型：
    - **test_case**: 普通测试用例，需要提供 test_case_steps
    - **test_case_bdd**: BDD 测试用例，需要提供 feature 和 scenario

    请求体字段：
    - **name**: 测试用例名称（必填）
    - **template**: 模板类型（可选，默认 test_case）
    - **description**: 描述（可选，支持 HTML）
    - **preconditions**: 前置条件（可选，支持 HTML）
    - **priority**: 优先级（可选）
    - **status**: 状态（可选）
    - **case_type**: 测试类型（可选）
    - **owner**: 负责人邮箱（可选）
    - **tags**: 标签列表（可选）
    - **issues**: 关联的 Jira issues（可选）
    - **custom_fields**: 自定义字段（可选）
    - **test_case_steps**: 测试步骤列表（普通测试用例）
    - **feature**: BDD Feature 描述（BDD 测试用例必填）
    - **scenario**: BDD Scenario 描述（BDD 测试用例必填）
    - **background**: BDD Background 描述（BDD 测试用例可选）
    """
    test_case = await service.create_test_case(
        project_identifier, data, current_user_id, folder_id
    )
    await db.commit()
    return SuccessResponse(success=True, data=test_case)

# ============ 更新测试用例接口 ============
# fmt: off  Mi80OmFIVnBZMlhscm9ua3VMazZhM2xhVUE9PTo4NTMxNTdhNQ==

@router.patch(
    "/test-cases/{test_case_identifier}",
    response_model=SuccessResponse[TestCaseInfo],
    summary="更新测试用例",
    description="更新指定测试用例的信息",
)
async def update_test_case(
    project_identifier: str,
    test_case_identifier: str,
    data: TestCaseUpdate,
    service: TestCaseServiceDep,
    db: DbSessionDep,
) -> SuccessResponse[TestCaseInfo]:
    """更新测试用例"""
    test_case = await service.update_test_case(
        project_identifier, test_case_identifier, data
    )
    await db.commit()
    return SuccessResponse(success=True, data=test_case)

# ============ 删除测试用例接口 ============

@router.delete(
    "/test-cases/{test_case_identifier}",
    response_model=MessageResponse,
    summary="删除测试用例",
    description="删除指定的测试用例",
)
async def delete_test_case(
    project_identifier: str,
    test_case_identifier: str,
    service: TestCaseServiceDep,
    db: DbSessionDep,
) -> MessageResponse:
    """删除测试用例"""
    message = await service.delete_test_case(project_identifier, test_case_identifier)
    await db.commit()
    return MessageResponse(success=True, message=message)

# ============ 批量操作接口 ============

@router.patch(
    "/test-cases",
    response_model=BulkOperationResponse,
    summary="批量更新测试用例",
    description="批量更新多个测试用例的信息",
)
async def bulk_update_test_cases(
    project_identifier: str,
    data: BulkTestCaseRequest,
    service: TestCaseServiceDep,
    db: DbSessionDep,
) -> BulkOperationResponse:
    """
    批量更新测试用例

    - **test_case_ids**: 要更新的测试用例标识符列表
    - **update_data**: 要更新的字段
    """
    affected_count = await service.bulk_update_test_cases(
        project_identifier,
        data.test_case_ids,
        data.update_data,
    )
    await db.commit()
    return BulkOperationResponse(
        success=True,
        message=f"成功更新 {affected_count} 个测试用例",
        affected_count=affected_count
    )

@router.patch(
    "/test-cases/with-operations",
    response_model=BulkOperationResponse,
    summary="带操作符的批量更新",
    description="使用操作符（ignore, replace, add, remove）批量更新测试用例",
)
async def bulk_update_with_operations(
    project_identifier: str,
    data: BulkEditWithOperationsRequest,
    service: TestCaseServiceDep,
    db: DbSessionDep,
) -> BulkOperationResponse:
    """
    带操作符的批量更新测试用例

    支持的操作符：
    - **ignore**: 保持现有值不变
    - **replace**: 用提供的值覆盖当前值
    - **add**: 将提供的值追加到现有列表（多值字段）
    - **remove**: 从现有列表中移除指定的值（多值字段）

    各字段支持的操作符：
    - automation_status, case_type, priority, state, owner, preconditions: ignore, replace
    - tags, issues, custom_fields: ignore, add, remove, replace
    """
    affected_count = await service.bulk_update_with_operations(
        project_identifier, data
    )
    await db.commit()
    return BulkOperationResponse(
        success=True,
        message=f"成功更新 {affected_count} 个测试用例",
        affected_count=affected_count
    )

@router.delete(
    "/test-cases",
    response_model=BulkOperationResponse,
    summary="批量删除测试用例",
    description="批量删除多个测试用例",
)
async def bulk_delete_test_cases(
    project_identifier: str,
    service: TestCaseServiceDep,
    db: DbSessionDep,
    data: BulkDeleteRequest = Body(...),
) -> BulkOperationResponse:
    """
    批量删除测试用例

    - **test_case_ids**: 要删除的测试用例标识符列表
    """
    affected_count = await service.bulk_delete_test_cases(
        project_identifier,
        data.test_case_ids,
    )
    await db.commit()
    return BulkOperationResponse(
        success=True,
        message=f"成功删除 {affected_count} 个测试用例",
        affected_count=affected_count
    )

# ============ BDD 导出接口 ============

@router.post(
    "/test-cases/export-bdd",
    response_model=ExportBDDResponse,
    summary="导出 BDD 测试用例",
    description="启动 BDD 测试用例导出任务，生成 .feature 文件",
)
async def export_bdd_test_cases(
    project_identifier: str,
    data: ExportBDDRequest,
    service: TestCaseServiceDep,
    export_service: ExportServiceDep,
) -> ExportBDDResponse:
    """
    导出 BDD 测试用例

    - **test_case_ids**: 要导出的测试用例标识符列表
    - **combine_into_one**: 是否合并为单个 .feature 文件
    - **combined_feature**: 合并后的 Feature 名称（combine_into_one=true 时必填）
    - **combined_background**: 合并后的 Background 内容（可选）

    返回导出任务 ID 和状态查询 URL
    """
    export_result = await export_service.start_bdd_export(
        project_identifier, data
    )
    return export_result

# ============ 测试用例历史接口 ============

@router.get(
    "/test-cases/{test_case_identifier}/history",
    response_model=TestCaseHistoryResponse,
    summary="获取测试用例历史",
    description="获取测试用例的变更历史记录",
)
async def get_test_case_history(
    project_identifier: str,
    test_case_identifier: str,
    service: TestCaseServiceDep,
    pagination: PaginationDep,
) -> TestCaseHistoryResponse:
    """
    获取测试用例历史

    返回测试用例的所有变更记录，包括：
    - 修改的字段
    - 修改前后的值
    - 修改时间
    - 修改人
    """
    history_data = await service.get_test_case_history(
        project_identifier,
        test_case_identifier,
        pagination.page,
        pagination.page_size
    )
    return history_data

# ============ 导出状态和下载接口（独立路由器） ============

exports_router = APIRouter(prefix="/exports")

@exports_router.get(
    "/{export_id}/status",
    response_model=ExportStatusResponse,
    summary="获取导出状态",
    description="获取 BDD 测试用例导出任务的状态",
)
async def get_export_status(
    export_id: str,
    export_service: ExportServiceDep,
) -> ExportStatusResponse:
    """
    获取导出状态

    返回导出任务的当前状态和下载 URL（如果已完成）
    """
    return await export_service.get_export_status(export_id)

# fmt: off  My80OmFIVnBZMlhscm9ua3VMazZhM2xhVUE9PTo4NTMxNTdhNQ==

@exports_router.get(
    "/{export_id}/download",
    summary="下载导出文件",
    description="下载已完成的 BDD 测试用例导出文件",
)
async def download_export(
    export_id: str,
    export_service: ExportServiceDep,
) -> StreamingResponse:
    """
    下载导出文件

    返回 .feature 文件或 .zip 压缩包
    """
    file_content, filename, content_type = await export_service.download_export(export_id)
    return StreamingResponse(
        iter([file_content]),
        media_type=content_type,
        headers={"Content-Disposition": build_content_disposition(filename)}
    )


# ============ 需求分析报告接口 ============

def _safe_filename(name: str) -> str:
    """将任意字符串转换为安全的文件名片段（与 requirement_tools.py 保持一致）。"""
    import re
    name = name.strip().replace(" ", "_")
    name = re.sub(r"[^\w\-一-鿿.]", "", name)
    return name[:80] or "requirement"


def _get_report_search_dirs(project_identifier: str) -> list[Path]:
    """获取需求分析报告可能存放的目录列表（按优先级排序）

    由于不同 Agent 运行时的 cwd 不同，且 AI 可能使用 write_file 直接写入
    而非 save_requirement_analysis 工具，文件可能位于多个不同路径。
    本函数返回所有可能的搜索路径，包括：
    1. 带 project_identifier 子目录的标准路径
    2. 不带子目录的扁平路径（AI 实际使用的路径）
    3. Agent 不同 cwd 下的解析路径
    """
    dirs = []
    seen = set()

    def _add(path: Path) -> None:
        """添加目录到列表，避免重复"""
        try:
            resolved = path.resolve()
            if resolved not in seen:
                seen.add(resolved)
                dirs.append(path)
        except (OSError, ValueError):
            pass

    # 推导项目根目录（从当前文件位置）
    # test_cases.py 在 backend/app/api/v2/，向上 4 层到项目根目录
    # backend/app/api/v2/ → backend/app/api/ → backend/app/ → backend/ → 项目根目录
    try:
        project_root = Path(__file__).resolve().parents[4]
    except IndexError:
        project_root = Path.cwd()

    # === 主路径：settings.api_workspace_root 解析的绝对路径 ===
    api_root = Path(settings.api_workspace_root).resolve()
    _add(api_root / "reports" / "requirement-analysis" / _safe_filename(project_identifier))
    _add(api_root / "reports" / "requirement-analysis")  # 扁平路径

    # === 兼容路径 1：从项目根目录推导的标准路径 ===
    # 当 settings.api_workspace_root 是相对路径时，从项目根目录解析
    if not Path(settings.api_workspace_root).is_absolute():
        standard_api = project_root / settings.api_workspace_root
        _add(standard_api / "reports" / "requirement-analysis" / _safe_filename(project_identifier))
        _add(standard_api / "reports" / "requirement-analysis")  # 扁平路径

    # === 兼容路径 2：testcase agent 运行时路径 ===
    # AI Agent 在 testcase workspace 下运行时，文件可能保存到此处
    testcase_base = project_root / "backend" / "workspace" / "testcase" / "workspace" / "api"
    _add(testcase_base / "reports" / "requirement-analysis" / _safe_filename(project_identifier))
    _add(testcase_base / "reports" / "requirement-analysis")  # 扁平路径

    # === 兼容路径 3：从当前文件位置推导 ===
    # 作为后备，从当前文件所在位置推导
    file_based_backend = Path(__file__).resolve().parents[3]  # 到 backend/
    file_based_root = file_based_backend.parent
    file_based_api = file_based_root / "backend" / "workspace" / "api"
    _add(file_based_api / "reports" / "requirement-analysis" / _safe_filename(project_identifier))
    _add(file_based_api / "reports" / "requirement-analysis")  # 扁平路径

    # === 兼容路径 4：从 cwd 推导（作为后备）===
    cwd = Path.cwd()
    cwd_api = cwd / "backend" / "workspace" / "api"
    _add(cwd_api / "reports" / "requirement-analysis" / _safe_filename(project_identifier))
    _add(cwd_api / "reports" / "requirement-analysis")  # 扁平路径

    return dirs


@router.get(
    "/requirement-reports",
    summary="获取需求分析报告列表",
    description="扫描 workspace 目录获取当前项目已生成的需求分析报告（Markdown 文件）列表",
)
async def get_requirement_reports(
    project_identifier: str,
) -> SuccessResponse:
    """
    获取需求分析报告列表

    扫描 api_workspace_root/reports/requirement-analysis/{project_identifier}/ 目录，
    返回所有 .md 文件的信息列表
    """
    search_dirs = _get_report_search_dirs(project_identifier)

    reports = []
    seen_files = set()  # 去重

    # 项目名称关键词（用于扁平路径过滤）
    project_keywords = _safe_filename(project_identifier).lower().replace("_", " ")

    for report_dir in search_dirs:
        if not report_dir.exists():
            continue
        for file_path in sorted(report_dir.glob("*.md"), key=lambda p: p.stat().st_mtime, reverse=True):
            if file_path.name in seen_files:
                continue
            seen_files.add(file_path.name)

            # 对于扁平路径（非 project_identifier 子目录），
            # 需要检查文件名是否包含项目名称关键词
            is_flat_dir = report_dir.name == "requirement-analysis"
            if is_flat_dir:
                # 从文件名提取项目标识（去掉时间戳前缀）
                file_stem = file_path.stem
                # 去掉时间戳前缀（如 20260718_2258_）
                import re
                clean_stem = re.sub(r'^\d{8}_\d{6}_', '', file_stem)
                # 检查文件名是否包含项目关键词
                if project_keywords and project_keywords not in clean_stem.lower():
                    # 也检查文件头部的 project 元数据
                    try:
                        header = file_path.read_text(encoding="utf-8", errors="ignore")[:500]
                        if f'project: {project_identifier}' not in header:
                            continue  # 跳过不相关的文件
                    except Exception:
                        continue  # 读取失败，跳过

            stat = file_path.stat()
            # 尝试从文件头部读取标题
            title = file_path.stem
            try:
                content = file_path.read_text(encoding="utf-8", errors="ignore")
                # 查找第一个 # 开头的标题行
                for line in content.split("\n")[:20]:
                    line = line.strip()
                    if line.startswith("# "):
                        title = line[2:].strip()
                        break
            except OSError as e:
                logger.warning(f"读取报告文件失败 {file_path}: {str(e)}")
            except UnicodeDecodeError as e:
                logger.warning(f"报告文件编码错误 {file_path}: {str(e)}")

            reports.append({
                "id": file_path.name,
                "filename": file_path.name,
                "title": title,
                "path": str(file_path),
                "size": stat.st_size,
                "created_at": datetime.fromtimestamp(stat.st_mtime).isoformat(),
            })

    # 按创建时间倒序排序
    reports.sort(key=lambda r: r["created_at"], reverse=True)

    return SuccessResponse(
        success=True,
        data=reports,
        message=f"找到 {len(reports)} 个需求分析报告",
    )


@router.post(
    "/requirement-reports/{filename}/export",
    summary="导出需求分析报告",
    description="导出指定的需求分析报告 Markdown 文件",
)
async def export_requirement_report(
    project_identifier: str,
    filename: str,
) -> StreamingResponse:
    """
    导出需求分析报告

    返回指定的 Markdown 文件内容
    """
    # 安全检查：防止路径遍历
    safe_filename = os.path.basename(filename)
    if not safe_filename.endswith(".md"):
        raise BadRequestException("仅支持导出 .md 文件")

    # 在所有可能的目录中查找文件
    search_dirs = _get_report_search_dirs(project_identifier)
    file_path = None
    for report_dir in search_dirs:
        candidate = report_dir / safe_filename
        if candidate.exists():
            file_path = candidate
            break

    if not file_path:
        raise BadRequestException(f"文件不存在: {safe_filename}")

    # 确保文件在允许的目录内（安全检查）
    found = False
    for report_dir in search_dirs:
        try:
            file_path.resolve().relative_to(report_dir.resolve())
            found = True
            break
        except ValueError:
            continue

    if not found:
        logger.warning(f"检测到非法的文件路径访问尝试: {filename}")
        raise BadRequestException("非法的文件路径")

    content = file_path.read_text(encoding="utf-8")

    return StreamingResponse(
        iter([content.encode("utf-8")]),
        media_type="text/markdown; charset=utf-8",
        headers={"Content-Disposition": build_content_disposition(safe_filename)}
    )


# 用例类型中文映射
_CASE_TYPE_LABELS = {
    "functional": "功能",
    "smoke_sanity": "冒烟",
    "regression": "回归",
    "security": "安全",
    "performance": "性能",
    "usability": "可用性",
    "acceptance": "验收",
    "compatibility": "兼容",
    "integration": "集成",
    "exploratory": "探索",
    "other": "其他",
}


def _strip_html(text: str) -> str:
    """移除 HTML 标签"""
    if not text:
        return ""
    import re
    text = re.sub(r'<p>', '', text)
    text = re.sub(r'</p>', '\n', text)
    text = re.sub(r'<br\s*/?>', '\n', text)
    text = re.sub(r'<[^>]+>', '', text)
    return text.strip()


def _clean_title(name: str, identifier: str, case_type: str, keyword: str) -> str:
    """
    清理用例标题，去掉编号和类型标记，返回纯标题

    系统提示词要求用例标题不包含编号和类型信息
    """
    if not name:
        return ""

    import re

    title = name

    # 去掉开头的编号前缀（如 TC-084、TC-ADM-001 等）
    # 匹配模式：TC-xxx、TC-xxx-xxx、TC-xxx: 等
    title = re.sub(r'^TC-[A-Z0-9-]+[:\s]*', '', title)
    title = re.sub(r'^TC-\d+[:\s]*', '', title)

    # 去掉末尾的类型标记（如（边界）、（正向）、（反向）等）
    title = re.sub(r'[\(（](边界|正向|反向|异常|正常)[\)）]$', '', title)

    # 去掉末尾的优先级标记（如（高优先级）、（中优先级）、（低优先级）等）
    title = re.sub(r'[\(（](高|中|低)优先级[\)）]$', '', title)

    # 去掉末尾的模块/类型组合标记（如（行政组织 / 异常类测试 / 高优先级））
    title = re.sub(r'[\(（][^\(（]*[/／][^\(（]*[\)）]$', '', title)

    return title.strip()
