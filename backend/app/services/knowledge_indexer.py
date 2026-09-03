"""
知识索引服务

负责文档切片、embedding 生成和入库
"""

import logging
from typing import Optional
from uuid import UUID

from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select

from app.models.knowledge_document import KnowledgeDocument, DocumentStatus
from app.models.knowledge_chunk import KnowledgeChunk
from app.config.settings import settings

logger = logging.getLogger(__name__)


class KnowledgeIndexer:
    """
    知识索引器

    负责文档切片、生成 embedding、入库
    """

    def __init__(self, session: AsyncSession):
        self.session = session
        self.chunk_size = settings.rag_chunk_size
        self.chunk_overlap = settings.rag_chunk_overlap

    async def index_document(self, document_id: UUID) -> bool:
        """
        对文档执行完整的索引流程：下载 -> 提取文本 -> 切片 -> 生成 embedding -> 入库

        Args:
            document_id: 文档 ID

        Returns:
            bool: 是否成功
        """
        start_time = __import__('time').time()

        try:
            # 获取文档信息
            result = await self.session.execute(
                select(KnowledgeDocument).where(KnowledgeDocument.id == document_id)
            )
            document = result.scalar_one_or_none()
            if not document:
                logger.error(f"文档不存在: {document_id}")
                return False

            # 更新状态为处理中
            document.status = DocumentStatus.PROCESSING
            await self.session.flush()

            # 从 MinIO 下载文件
            text_content = await self._download_and_extract_text(document)
            if not text_content:
                document.status = DocumentStatus.FAILED
                document.error_message = "无法提取文档文本内容"
                await self.session.flush()
                return False

            # 切片
            chunks = self.chunk_document(text_content)
            logger.info(f"文档 {document_id} 切片完成，共 {len(chunks)} 个切片")

            # 生成 embedding 并入库
            embeddings = await self.generate_embeddings(chunks)

            # 保存切片
            for i, (chunk_text, embedding) in enumerate(zip(chunks, embeddings)):
                chunk = KnowledgeChunk(
                    document_id=document_id,
                    space_id=document.space_id,
                    content=chunk_text,
                    embedding=embedding,
                    chunk_index=i,
                    meta_data={"source": document.file_name},
                )
                self.session.add(chunk)

            # 更新文档状态
            document.status = DocumentStatus.INDEXED
            document.chunk_count = len(chunks)
            await self.session.flush()

            elapsed = __import__('time').time() - start_time
            logger.info(f"文档 {document_id} 索引完成，耗时 {elapsed:.2f}s")
            return True

        except Exception as e:
            logger.error(f"文档索引失败 {document_id}: {e}", exc_info=True)
            # 更新失败状态
            try:
                result = await self.session.execute(
                    select(KnowledgeDocument).where(KnowledgeDocument.id == document_id)
                )
                document = result.scalar_one_or_none()
                if document:
                    document.status = DocumentStatus.FAILED
                    document.error_message = str(e)[:500]
                    await self.session.flush()
            except Exception:
                pass
            return False

    async def _download_and_extract_text(self, document: KnowledgeDocument) -> Optional[str]:
        """从 MinIO 下载文件并提取文本"""
        try:
            from app.config.minio_client import MinIOClient

            data = MinIOClient.download_file(document.file_path)
            if not data:
                return None

            # 根据文件类型提取文本
            content_type = document.content_type
            file_name = document.file_name.lower()

            if content_type == "application/pdf" or file_name.endswith(".pdf"):
                return self._extract_pdf_text(data)
            elif content_type in (
                "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
                "application/msword",
            ) or file_name.endswith(".docx"):
                return self._extract_word_text(data)
            elif content_type in (
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                "application/vnd.ms-excel",
            ) or file_name.endswith(".xlsx"):
                return self._extract_excel_text(data)
            elif content_type == "text/plain" or file_name.endswith(".txt"):
                return data.decode("utf-8", errors="ignore")
            else:
                # 尝试作为文本处理
                return data.decode("utf-8", errors="ignore")

        except Exception as e:
            logger.error(f"提取文档文本失败: {e}")
            return None

    def _extract_pdf_text(self, data: bytes) -> Optional[str]:
        """从 PDF 提取文本"""
        try:
            import fitz  # PyMuPDF
            doc = fitz.open(stream=data, filetype="pdf")
            texts = []
            for page in doc:
                texts.append(page.get_text("text"))
            return "\n".join(texts)
        except ImportError:
            logger.warning("PyMuPDF 未安装，PDF 提取失败")
            return None
        except Exception as e:
            logger.error(f"PDF 文本提取失败: {e}")
            return None

    def _extract_word_text(self, data: bytes) -> Optional[str]:
        """从 Word 文档提取文本"""
        try:
            from docx import Document
            from io import BytesIO

            doc = Document(BytesIO(data))
            paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
            return "\n".join(paragraphs)
        except ImportError:
            logger.warning("python-docx 未安装，Word 提取失败")
            return None
        except Exception as e:
            logger.error(f"Word 文本提取失败: {e}")
            return None

    def _extract_excel_text(self, data: bytes) -> Optional[str]:
        """从 Excel 文档提取文本（所有 sheet 的单元格内容）"""
        try:
            import openpyxl
            from io import BytesIO

            wb = openpyxl.load_workbook(BytesIO(data), data_only=True)
            texts = []
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                texts.append(f"=== Sheet: {sheet_name} ===")
                for row in sheet.iter_rows():
                    row_texts = []
                    for cell in row:
                        if cell.value is not None:
                            row_texts.append(str(cell.value))
                    if row_texts:
                        texts.append(" | ".join(row_texts))
            return "\n".join(texts)
        except ImportError:
            logger.warning("openpyxl 未安装，Excel 提取失败")
            return None
        except Exception as e:
            logger.error(f"Excel 文本提取失败: {e}")
            return None

    def chunk_document(self, text: str) -> list[str]:
        """
        按固定长度+重叠策略切片文档

        Args:
            text: 文档全文

        Returns:
            list[str]: 切片列表
        """
        if not text:
            return []

        chunks = []
        start = 0
        text_len = len(text)

        while start < text_len:
            end = min(start + self.chunk_size, text_len)
            chunk = text[start:end]
            chunks.append(chunk)
            start += self.chunk_size - self.chunk_overlap

        return chunks

    async def generate_embeddings(self, chunks: list[str]) -> list[str]:
        """
        为文本切片生成 embedding 向量

        Args:
            chunks: 文本切片列表

        Returns:
            list[str]: embedding 向量列表（逗号分隔的浮点数字符串）
        """
        if not chunks:
            return []

        # 检查是否有 embedding API 配置（与 anything-chat-rag 框架保持一致）
        # 优先使用 embedding_binding_api_key，其次复用 deepseek_api_key
        api_key = getattr(settings, 'embedding_binding_api_key', None) or getattr(settings, 'deepseek_api_key', None)

        if not api_key:
            logger.warning("未配置 Embedding API Key（embedding_binding_api_key 或 deepseek_api_key），使用零向量占位")
            dimensions = getattr(settings, 'embedding_dim', 1024)
            return [",".join(["0.0"] * dimensions) for _ in chunks]

        try:
            import httpx

            embeddings = []
            # 使用与 anything-chat-rag 框架一致的模型配置
            model = getattr(settings, 'embedding_model', 'Qwen/Qwen3-Embedding-0.6B')
            api_base = getattr(settings, 'embedding_binding_host', 'https://llmapi.dtyunxi.cn/v1')

            async with httpx.AsyncClient() as client:
                for chunk in chunks:
                    response = await client.post(
                        f"{api_base}/embeddings",
                        headers={
                            "Authorization": f"Bearer {api_key}",
                            "Content-Type": "application/json",
                        },
                        json={
                            "model": model,
                            "input": chunk[:settings.embedding_token_limit],  # 使用配置的 token 限制
                        },
                        timeout=60.0,
                    )
                    response.raise_for_status()
                    data = response.json()
                    vector = data["data"][0]["embedding"]
                    embeddings.append(",".join(map(str, vector)))

            return embeddings

        except Exception as e:
            logger.error(f"Embedding 生成失败: {e}")
            # 降级：返回零向量
            dimensions = getattr(settings, 'embedding_dim', 1024)
            return [",".join(["0.0"] * dimensions) for _ in chunks]

    async def reindex_space(self, space_id: UUID) -> dict:
        """
        重建整个知识空间的索引

        Args:
            space_id: 知识空间 ID

        Returns:
            dict: 重建结果统计
        """
        # 删除旧切片
        from sqlalchemy import delete
        await self.session.execute(
            delete(KnowledgeChunk).where(KnowledgeChunk.space_id == space_id)
        )

        # 获取空间下所有文档
        result = await self.session.execute(
            select(KnowledgeDocument).where(
                KnowledgeDocument.space_id == space_id
            )
        )
        documents = result.scalars().all()

        success_count = 0
        failed_count = 0

        for doc in documents:
            doc.status = DocumentStatus.PENDING
            doc.chunk_count = 0
            doc.error_message = None

        await self.session.flush()

        # 重新索引每个文档
        for doc in documents:
            if await self.index_document(doc.id):
                success_count += 1
            else:
                failed_count += 1

        return {
            "total": len(documents),
            "success": success_count,
            "failed": failed_count,
        }
